#!/usr/bin/env python
# -*- coding: utf-8 -*-

#po_フォルダからデータを読みこんで、DBに書き込み

import openpyxl 
import datetime
import sqlite3
import glob
import datetime

SFILE = "../../tfc_sql/tfc.sqlite"
POFILES="../../tfcpos/*.xlsx"

ITEMROWBEGIN = 16 #アイテム行は１から数えて何番目から
ITEMCOLUMNBEGIN = 2 #アイテム列は、A=１から数えて何番目
PODATE ="J6"
PON ="J7"


class ReadPos:
    def __init__(self):
        pofiles = glob.glob(POFILES)
        for pof in pofiles:
            self.insert_po(pof)
        #self.insert_po(pofiles[4])

    def insert_po(self, filename):
        con=sqlite3.connect(SFILE)
        cur = con.cursor()
        #とりあえず、ponとpodをINSERTしてidを取得
        sql1 = "INSERT INTO po(pon, pod) VALUES(?, ?)"
        book = openpyxl.load_workbook(filename)
        sheet = book['PO']
        pod = sheet[PODATE].value.strftime("%Y-%m-%d")
        pon = sheet[PON].value
        print("PO NO:", pon)
        cur.execute(sql1, (pon, pod))
        con.commit()
        poid = cur.lastrowid

        sql2 = 'SELECT id FROM tfc_code where item = ?'
        sql3 = 'INSERT INTO poline (code_id, qty, remark, om, po_id) VALUES(?, ?, ?, ?, ?)'

        r = ITEMROWBEGIN
        while( True ):
            c = ITEMCOLUMNBEGIN
            item = sheet.cell(row=r, column=c).value
            if item == None :
                break
            else:
                if item.startswith('013'):
                    item = item.replace("013","")

                cur.execute(sql2, (item,))
                codeid = cur.fetchone()
                if codeid == None :
                    print("PO No.{} の {} 行目 {} はコードがないのでスキップします。".format(pon, r, item)) 
                    r+=1
                    continue
                else:
                    print("Writing..", item)
                    remarks = sheet.cell(row=r, column=c+2).value
                    qty = sheet.cell(row=r, column=c+3).value
                    om = sheet.cell(row=r, column=c+10).value 
                    cur.execute(sql3, (codeid[0], qty, remarks, om, poid))
                    con.commit()
                    r+=1

        sql4 = "UPDATE po SET per=?, port=?, shipto=?, comment=?, ETD=? where id =?"
        #shipment per の値取得
        per = sheet.cell(row=r+3, column=3).value
        #shipto の値取得
        shipto = sheet.cell(row=r+4, column=3).value
        #port の値取得
        port = sheet.cell(row=r+5, column=3).value
        #comment 取得
        comment = sheet.cell(row=r+2, column=3).value
        #ETD 取得
        etd =  sheet.cell(row=r+3, column=5).value
        if type(etd)  != datetime.datetime:
            etd =  pod
        else:
            etd =  sheet.cell(row=r+3, column=5).value.strftime("%Y-%m-%d")

        cur.execute(sql4, (per, port, shipto, comment, etd, poid))
        con.commit()

        con.close()

        print('r=', r)


r = ReadPos()
