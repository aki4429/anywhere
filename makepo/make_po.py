#!/usr/bin/env python
# -*- coding: utf-8 -*-

#po_linesファイルからデータを読みこんで、PO作成。
#発注データを保存する => データベースなど

import os
import csv 
import openpyxl 
import datetime
import poline
import condition
import po
import sqlite3

#file_name = tools.get_latest_modified_file_path(os.curdir, "*xl*")
POFILE = "po.xlsx"
POLINES = "../po_lines_keep.csv"
CONDFILE = "conditions.csv"
PONOFILE = "pono.txt"
TFCS = "../tfc.sqlite"

ITEMROWBEGIN = 16 #アイテム行は１から数えて何番目から
ITEMCOLUMNBEGIN = 2 #アイテム列は、A=１から数えて何番目

PODATE ="J6"

MAX = 300 #ベースPOのアイテム行の最後行 index

class MakePo:
    def __init__(self):
        pd = self.sel_import(CONDFILE)
        self.po_data = self.read_lines(POLINES)
        self.show_podata()
        self.read_write_po(POFILE, self.po_data, pd)
        self.write_posql(pd, self.po_data)
        #self.set_conditions()
        #self.make_po()
        #self.write_data()

    #エクセルPOファイル雛形を読み込みデータを書き込む
    def read_write_po(self, filename, po_data, pd):
        book = openpyxl.load_workbook(filename)
        sheet = book['PO']
        r = ITEMROWBEGIN
        for pl in po_data:
            c = ITEMCOLUMNBEGIN
            sheet.cell(row=r, column=c).value = pl.item
            sheet.cell(row=r, column=c+1).value = pl.description
            sheet.cell(row=r, column=c+2).value = pl.remarks
            sheet.cell(row=r, column=c+3).value = int(float(pl.qty))
            sheet.cell(row=r, column=c+4).value = pl.unit
            sheet.cell(row=r, column=c+5).value = (float(pl.u_price) 
                    if pl.u_price.replace(".", "").isdigit() else 0)
            #金額計算式挿入
            sheet.cell(row=r, column=c+6).value = "=E{0}*G{0}".format(r)
            sheet.cell(row=r, column=c+8).value = pl.our_item
            #容積計算式挿入
            sheet.cell(row=r, column=c+9).value = "=E{0}*N{0}".format(r)
            sheet.cell(row=r, column=c+10).value = pl.om
            sheet.cell(row=r, column=c+12).value = (float(pl.u_M3) 
                    if pl.u_M3.replace(".", "").isdigit() else 0)
            r += 1

        #書き込み最終 index から MAX 行-1行を削除
        sheet.delete_rows(r+1, MAX-r )
        #数量合計式挿入
        sheet.cell(row=r+1, column=c+3).value = "=SUM(E{0}:E{1})".format(ITEMROWBEGIN, r) 
        #金額合計式挿入
        sheet.cell(row=r+1, column=c+6).value = "=SUM(H{0}:H{1})".format(ITEMROWBEGIN, r) 
        #容積合計式挿入
        sheet.cell(row=r+1, column=c+9).value = "=SUM(K{0}:K{1})".format(ITEMROWBEGIN, r) 
        sheet.merge_cells("E{0}:H{0}".format( r+3 ))

        #発注日挿入
        sheet.cell(row=6, column=10).value = pd.podate

        #comment 挿入
        sheet.cell(row=r+2, column=3).value = pd.condition.comment

        #PONO挿入
        sheet.cell(row=7, column=10).value = pd.pono

        #shipment per挿入
        sheet.cell(row=r+3, column=2).value = "Shipment per:"
        sheet.cell(row=r+3, column=3).value = pd.condition.shipment_per

        #ship to挿入
        sheet.cell(row=r+4, column=2).value = "Ship to:"
        sheet.cell(row=r+4, column=3).value = pd.condition.shipto_1

        #ship to 2 なしの場合
        if not pd.condition.shipto_2 :
            #Via タイトル : データ挿入
            sheet.cell(row=r+5, column=2).value = "Via:"
            sheet.cell(row=r+5, column=3).value = pd.condition.via
            #forwarder タイトル : データ挿入
            sheet.cell(row=r+6, column=2).value = "Forwarder: "
            sheet.cell(row=r+6, column=3).value = pd.condition.forwarder
            
            #Delivery タイトル : データ挿入
            sheet.cell(row=r+3, column=4).value = "Delivery(ETD): "
            sheet.cell(row=r+3, column=5).value = pd.etddate
            #Tradeterm タイトル : データ挿入
            sheet.cell(row=r+4, column=4).value = "Trade Term:: "
            sheet.cell(row=r+4, column=5).value = pd.condition.trade_term
            #Payment タイトル : データ挿入
            sheet.cell(row=r+5, column=4).value = "Payment: "
            sheet.cell(row=r+5, column=5).value = pd.condition.payment
            #Insurance タイトル : データ挿入
            sheet.cell(row=r+6, column=4).value = "Insurance: "
            sheet.cell(row=r+6, column=5).value = pd.condition.insurance

            # サイン欄までの空白行を2行削除
            sheet.delete_rows(r+7, 2 )

        else:
            #ship to_2,3,4, 5挿入
            sheet.cell(row=r+5, column=3).value = pd.condition.shipto_2
            sheet.cell(row=r+6, column=3).value = pd.condition.shipto_3
            sheet.cell(row=r+7, column=3).value = pd.condition.shipto_4
            sheet.cell(row=r+8, column=3).value = pd.condition.shipto_5

            #Delivery タイトル : データ挿入
            sheet.cell(row=r+3, column=4).value = "Delivery(ETD): "
            sheet.cell(row=r+3, column=5).value = pd.etddate

            #Tradeterm タイトル : データ挿入
            sheet.cell(row=r+4, column=4).value = "Trade Term:: "
            sheet.cell(row=r+4, column=5).value = pd.condition.trade_term

            #Payment タイトル : データ挿入
            sheet.cell(row=r+5, column=4).value = "Payment: "
            sheet.cell(row=r+5, column=5).value = pd.condition.payment

            #Insurance タイトル : データ挿入
            sheet.cell(row=r+6, column=4).value = "Insurance: "
            sheet.cell(row=r+6, column=5).value = pd.condition.insurance
            

        pofilename = "PO" + pd.pono + pd.etddate.strftime("(%m%d_") +  pd.condition.shipment_per + pd.condition.via +").xlsx"
        pofilename = pofilename.strip().replace(" ","_")

        book.save(pofilename)
        print(pofilename + "を書き込みました。")

    def write_posql(self, pd, po_data):
        con = sqlite3.connect(TFCS)
        cur = con.cursor()
        cur.execute('INSERT INTO po (pod, pon, per, port, shipto, comment, ETD) VALUES (?, ?, ?, ?, ?, ?, ?)',
                (pd.podate, pd.pono, pd.condition.shipment_per, pd.condition.via, pd.condition.shipto_1, pd.condition.comment, pd.etddate))
        con.commit()

        lastid = cur.lastrowid
        for pl in po_data:
            if pl.code_id == "":
                cur.execute("select id from tfc_code where item = ?",
                        (pl.item,))
                codeid = cur.fetchone()[0]
                #print("codeid1", codeid)
            else:
                codeid = pl.code_id
                #print("codeid2", codeid)

            cur.execute('INSERT INTO poline (code_id, qty, remark, om, balance, po_id) VALUES(?, ?, ?, ?, ?, ?)', (codeid, pl.qty, pl.remarks, pl.om, pl.qty, lastid ))
            print("Writing code_id, pl.item, pl.qty, pl.om", codeid, pl.qty, pl.remarks, pl.om)

        con.commit()
        con.close()


    #発注データpo_linesの読み込み
    def read_lines(self, filename):
        data = []
        with open(filename, 'r', encoding='CP932') as f:
            reader = csv.reader(f)
            for row in reader: 
                pl = poline.Poline(*row)
                data.append(pl)

        return data

    def show_podata(self):
        for pl in self.po_data:
            pl.show_line()

    #def set_conditions(self):
        

    #輸入条件を選びます
    def sel_import(self, filename):
        data = []
        #ファイルの輸入条件を読み込み
        with open(filename, 'r', encoding='CP932') as f:
            reader = csv.reader(f, delimiter= "|")
            headder = next(reader)
            for row in reader: 
                cd = condition.Condition(*row)
                data.append(cd)
  
        #輸入条件メニュー作成
        num=1
        for cd in data:
            print(str(num) + ")" + cd.name)
            num += 1

        #輸入条件選択
        ans = 'n'
        cond = "" 
        while ans != 'y':
            print("="*30)
            index = int(input("輸入条件を選んでください:"))
            data[index-1].show_line()
            print("-"*30)
            ans = input("この条件でよいですか(y/n):")
            if ans == 'y' :
                cond = data[index -1]
            
        cond.show()

        #PO発注日選択(今日かそれ以外)
        podate_menu = " 1) {0} \n 2) それ以外".format(datetime.date.today())
        print(podate_menu)
        ans = input("PO発注日を選択してください。:")
        if ans == "1" :
            podate = datetime.date.today()
        else:
            str_date = input("PO発注日を入力してください。(20190331):")
            podate = datetime.datetime.strptime(str_date, "%Y%m%d").date()

        #ETD設定
        str_date = input("ETD 出港日を入力してください。(20190331):")
        etddate = datetime.datetime.strptime(str_date, "%Y%m%d").date()

        #POナンバー設定
        with open(PONOFILE, "r") as f:
            nu = int(f.read()) + 1 

        ans = input("PO NO. を{0}に設定します。いいですか。(y/n)".format(nu))
        if ans == "y":
            pono = "H" + str( nu )
        else:
            nu = int(input("PO NO. を設定してください。({0})".format(nu)))
            pono = "H" + str( nu )

        #PONOファイルを更新します。
        with open(PONOFILE, "w") as w:
            w.write(str(nu))

        pd = po.Po(cond, podate, etddate, pono)
        return pd



m = MakePo()
