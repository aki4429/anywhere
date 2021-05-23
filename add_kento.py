# vim:fileencoding=utf-8

"""
検討表の在庫発注数量をkako_juchu.csv と
torikomi_juchu.csv　に追加する。
"""

import glob
import openpyxl
import csv
import set_counter

#検討表ファイル名取得
#頭に「検討表」がつく/保存フォルダは podata
kento_name = glob.glob('./podata/検討表*.xlsx')[-1]
print("検討表ファイル名は:", kento_name)
ans = input('ok?(y=書き込みます/n=中止:')
if ans == 'n':
    exit() 

#検討表を開きます。
wb = openpyxl.load_workbook(kento_name)
sheet = wb['kento']

#行列番号は1始まり、Row 5から、Col 2=コード、25=(Y列)今回発注
#発注辞書に保存
orders = {}

i=0
while((i+5) <=600):
    code = sheet.cell(row=i+5, column=2).value
    #qty = sheet.cell(row=i+5, column=26).value  #Z
    #qty = sheet.cell(row=i+5, column=25).value  #Y
    #qty = sheet.cell(row=i+5, column=24).value  #X
    qty = sheet.cell(row=i+5, column=22).value  #V
    #qty = sheet.cell(row=i+5, column=23).value  #W
    #print(type(qty), qty, not qty)

    if qty :
        #print('code;', code, 'qty', qty)
        orders[code] = qty

    i += 1

wb.close()

kako_data = [] #kako_juchu.csv用データ
torikomi_line = [] #torikomi_juchu.csv用データ格納

for k, v in orders.items():
    kako_line = [''] * 8 #kako_juchu.csv用データ行格納
    kako_line[0] = k
    kako_line[4] = v
    kako_data.append(kako_line)

#kako_juchu.csv 読み込み
with open('kako_juchu.csv', encoding='CP932') as f:
    reader = csv.reader(f)
    for row in reader:
        kako_data.append(row)

with open('kako_juchu_2.csv', 'w', encoding='CP932') as f:
    writer = csv.writer(f)
    writer.writerows(kako_data)

#変換用ファイルを読み込み
change = []
with open('changer.csv') as f:
    reader = csv.reader(f)
    for row in reader:
        change.append(row)

#set counter の開始
s = set_counter.SetCounter()


torikomi_data = [] #torikomi_juchu.csv用データ
for k, v in orders.items():
    torikomi_line = [''] * 6 #kako_juchu.csv用データ行格納
    torikomi_line[0] = k
    torikomi_line[4] = v
    s.set(k, v) #set counter で数えておきます。

    #検討表でobicコードが異なるものは変換する。
    for d in change:
        if d[0] == k:
            k= d[1]

    torikomi_line[5] = k.replace('013CH', '013')
    torikomi_line[5] = k.replace('232WI', '232W')
    torikomi_line[5] = k.replace('271I', '271')
    torikomi_data.append(torikomi_line)

count = s.count
for c, q in count.items():
    torikomi_line = [''] * 6 #kako_juchu.csv用データ行格納
    torikomi_line[0] = c
    torikomi_line[1] = 'セット品'
    torikomi_line[4] = q
    torikomi_line[5] = c
    torikomi_data.append(torikomi_line)


#print('tori', torikomi_data)
#torikomi_juchu.csv 読み込み
with open('torikomi_juchu.csv', encoding='CP932') as f:
    reader = csv.reader(f)
    for row in reader:
        torikomi_data.append(row)

with open('torikomi_juchu_2.csv', 'w', encoding='CP932') as f:
    writer = csv.writer(f)
    writer.writerows(torikomi_data)

