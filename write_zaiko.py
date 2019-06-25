#!/usr/bin/env python
# -*- coding: utf-8 -*-

#在庫報告を作成するプログラム。
#DBからzaiko　フラグのデータをダウンロードする。
#品名、カテゴリーの順番に

DBFILE='tfc.sqlite'

CAT_ORDER = {'布地':1,'ﾇｰﾄﾞ':2,'ｶﾊﾞｰ':3,'INCOON':4, 'INCOON BED':5, '脚':6,'ｸｯｼｮﾝ':7, 'HYPERFLEX':8, '旧モデル':9}

MENU ="""
作成する表を選んでください。
===========================
1) TFC在庫表
2) TFC検討表
--------------------------
番号で選んでください。(1/2/.. or q=終了): """

ZAIKOF = 'zaiko.csv'
KENTOF = 'kento.csv'
ZHYO = 'zaiko_hyo.csv'
KHYO = 'kento_hyo.csv'
KENTO_C = 'kentohyo.csv' #検討表のコード並び
ZEXCEL = 'TFC_zaiko.xlsx'
                
import zaiko_read
import make_yotei
import pandas as pd
from pandas import DataFrame, Series
import numpy as np
import openpyxl
import mhparse as mh
import datetime
import dateutil.parser
import os

import sqlite3

class WriteZaiko:
    def __init__(self):
        con = sqlite3.connect(DBFILE)
        cur = con.cursor()

        ans = ''
        while ans != 'q':
            ans = input(MENU)

            if ans == '1':
                print("在庫表を作成します")
                #input()
                #DB tfc_codeテーブルから在庫フラグがあるデータをゲット
                df = self.get_zaiko(con)
                zaiko_file_name = ZAIKOF
                hyo_file_name = ZHYO
                df = self.order_by_code(df, ans)
                ans = 'q'
            elif ans == '2':
                #DB tfc_codeから 検討フラグがあるデータをゲット
                print("検討表を作成します")
                df = self.get_kento(con)
                zaiko_file_name = KENTOF
                hyo_file_name = KHYO
                df = self.order_by_code(df, ans)
                ans = 'q'
            elif ans == 'q':
                os.system('clear')
                return None
            else:
                print("半角番号を選んでください。")
                input("return to continue")
                continue


        #print('df', df)
        #input()
        df.to_csv(zaiko_file_name)
        
        k = zaiko_read.ZaikoRead()
        while k.result == 0 :
            print("kentohyo に適正にファイルをアップロードしてください。")
            input()
            k = zaiko_read.ZaikoRead()

        hyo = df.join([k.df], how='left')
        hyo = hyo.reindex(["在庫", "受注", "cat"], axis=1)
        y = make_yotei.MakeYotei()
        hyo = hyo.join([y.frame], how='left')
        #print(hyo)
        if hyo_file_name == KHYO:
            kento_code = pd.read_csv(KENTO_C, index_col='品目CD')
            hyo = kento_code.join(hyo, how='left')
        elif hyo_file_name == ZHYO:
            self.write_excel(hyo, k.get_date(), y)


        #print('hyo_file_name',hyo_file_name)
        #input()
        hyo.to_csv(hyo_file_name, encoding='CP932')
        #hyo.to_csv("kento_hyo.csv")

    def show(self, data):
        for d in data:
            for row in d: 
                print(row[0], row[1])

    def get_zaiko(self, con):
        #在庫フラグがあるデータをゲット
        #cur.execute("select cat, hcode from tfc_code where zaiko=1")
        #data = cur.fetchall()
        #return data
        data = pd.read_sql("select hcode, cat from tfc_code where zaiko=1",\
                con, index_col = 'hcode')

        return data

    def get_kento(self, con):
        #検討フラグがあるデータをゲット
        data = pd.read_sql("select hcode, cat from tfc_code where kento=1",\
                con, index_col = 'hcode')

        return data

    def order_by_code(self, df, ans):
        #df DataFrame の index にしているコードをlistで抽出
        codes = list(df.index.values)
        #モデル名末尾のIを削除
        removed =[]
        for code in codes:
            removed.append(code.replace("I-", "-"))

        codes = removed
        models =[]
        items =[]
        fabs =[]

        for code in codes:
            #-で区切って左がモデル名
            models.append(code.split('-')[0])
            if len(code.split('-')) >=2: 
                #-の右があれば、スペースで区切って左がアイテム
                items.append(code.split('-')[1].split(' ')[0])
                flag = 0
                #-の右をスペースで区切って右があれば、
                if len(code.split('-')[1].split(' ')) >=2:
                    for fab in code.split('-')[1].split(' '):
                        if '/' in fab:
                            # / を含む項目があるときfabs に追加
                            fabs.append(fab)
                            flag = 1
                if flag == 0:
                    fabs.append('')
            else:
                items.append('')
                fabs.append('')
                
        df['model'] = models
        df['item'] = items
        df['fab'] = fabs

        #在庫表作成の場合、CAT_ORDER順に並べ替える
        #print("ans:", ans)
        #input()
        if ans == '1':
            catnums = []
            for catname in df.cat.values:
                catnums.append( CAT_ORDER[catname] )

            df['catn'] = catnums

            df.sort_values(['catn', 'model', 'fab', 'item'], inplace = True)
            df.drop(['catn', 'model', 'fab', 'item'], axis = 1, inplace = True)

        return df


    def write_excel(self, hyo, kijunbi, yotei):
        wb = openpyxl.load_workbook(ZEXCEL)
        sheet = wb['zaiko']
        sheet['B1'] = kijunbi
        #コード記入
        j=0
        i=4 #4行目からスタート
        while i < len(hyo):
            sheet.cell(row=i, column=1, value = hyo.index[j]) 
            i += 1
            j += 1
       
        #在庫記入
        j=0
        i=4 #4行目からスタート
        while i < len(hyo):
            sheet.cell(row=i, column=2, value = hyo.iloc[j,0]) 
            i += 1
            j += 1
       
        #受注記入
        j=0
        i=4 #4行目からスタート
        while i < len(hyo):
            sheet.cell(row=i, column=3, value = hyo.iloc[j,1] )
            i += 1
            j += 1
       
        #有効残記入
        j=0
        i=4 #4行目からスタート
        while i < len(hyo):
            sheet.cell(row=i, column=4, value = '=B{0}-C{0}'.format(i)) 
            i += 1
            j += 1
       
        #カテゴリー記入
        j=0
        i=4 #4行目からスタート
        while i < len(hyo):
            sheet.cell(row=i, column=5, value = hyo.iloc[j,2]) 
            i += 1
            j += 1
       
        #予定入荷数記入
        j=0
        i=4 #4行目からスタート
        while i < len(hyo):
            k = 3
            while k < len(hyo.columns) :
                sheet.cell(row=i, column=3+k, value = hyo.iloc[j,k]) 
                k += 1
            i += 1
            j += 1
       
        #南濃取り込み日記入
        gyo=3 #3行目に記入
        j=3 #index 3 以降予定データ
        #６列目からスタート = j+3
        while j < len(hyo.columns):
            sheet.cell(row=gyo, column=j+3, value = dateutil.parser.parse(hyo.columns[j])) 
            j += 1
       
        #ETD記入
        gyo=2 #2行目に記入
        i=0
        j=3 #index 3 以降予定データ
        #６列目からスタート = j+3
        while j < len(hyo.columns):
            sheet.cell(row=gyo, column=j+3, value = dateutil.parser.parse(yotei.etds[i])) 
            j += 1
            i += 1
       
        #PO#, INV#記入
        i=0
        gyo=1 #1行目に記入
        j=3 #index 3 以降予定データ
        #６列目からスタート = j+3
        while j < len(hyo.columns):
            sheet.cell(row=gyo, column=j+3, value = yotei.pos[i]) 
            j += 1
            i += 1
        
        save_file = 'zaikohyo/TFC_zaiko_{0}.xlsx'.format(mh.parse(kijunbi))
        wb.save(save_file)
        print("{}を保存しました。".format(save_file))
        








#w = WriteZaiko()
